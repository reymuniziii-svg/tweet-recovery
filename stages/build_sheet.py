"""Stage 3 — Report: jsonl results -> xlsx (Tweets/Failed/Summary) + CSVs.

Replicates the proven three-sheet template: verbatim tweets sorted by
date, an honest failure ledger, and a summary with method + caveats.
"""

import csv
import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

TWEET_COLS = [
    ("Date", 20), ("Tweet Text", 90), ("Date Confidence", 15), ("Source", 20),
    ("Archive URL", 70), ("Tweet ID", 22), ("Capture Timestamp", 18),
    ("Recovered Via", 14), ("Is Retweet", 11), ("Retweeted User", 18),
]
FAILED_COLS = [
    ("Tweet ID", 22), ("Reason", 32), ("Needs Browser", 14),
    ("Attempted Captures", 40), ("Original URL", 70),
]

METHOD_NOTE = (
    "Text is taken verbatim from archived snapshots (old tweet-text element or "
    "og:description). Nothing is fabricated — missing tweets stay missing. "
    "Dates are decoded from each tweet's ID (Twitter snowflake), validated to "
    "match the timestamps the archive captured. Failures are unrendered "
    "JavaScript app shells or captures the archive returns as 404. "
    "Archived profile-page captures are also mined for the ~20 tweets each "
    "shows; retweet rows carry the retweeted author's text verbatim and are "
    "dated by the retweet's own ID when available, else by capture date."
)


def _read_jsonl(path: Path):
    if not path.exists():
        return []
    with path.open() as f:
        return [json.loads(line) for line in f if line.strip()]


def build_sheet(handle: str, workdir: Path, outdir: Path, raw_captures: int, unique_ids: int,
                page_stats: dict | None = None, manifest_ids: set | None = None):
    recovered = sorted(_read_jsonl(workdir / "recovered.jsonl"), key=lambda r: r["date"])
    # A timeline pass can recover an ID that previously failed (JS shell
    # on the permalink, rendered on a feed page): recovered wins.
    recovered_ids = {r["tweet_id"] for r in recovered}
    failed = sorted(
        (r for r in _read_jsonl(workdir / "failed.jsonl")
         if r["tweet_id"] not in recovered_ids),
        key=lambda r: int(r["tweet_id"]),
    )
    pages_done = _read_jsonl(workdir / "pages_done.jsonl")
    outdir.mkdir(parents=True, exist_ok=True)

    # --- CSVs (agent- and Sheets-import-friendly) ---
    with (outdir / "tweets.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow([c for c, _ in TWEET_COLS])
        for r in recovered:
            w.writerow(_tweet_row(r))
    with (outdir / "failed.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow([c for c, _ in FAILED_COLS])
        for r in failed:
            w.writerow(_failed_row(r))

    # --- Workbook ---
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    _write_summary(ws, handle, recovered, failed, raw_captures, unique_ids,
                   page_stats, pages_done, manifest_ids)

    ws = wb.create_sheet("Tweets")
    _write_table(ws, TWEET_COLS, [_tweet_row(r) for r in recovered])

    ws = wb.create_sheet("Failed")
    _write_table(ws, FAILED_COLS, [_failed_row(r) for r in failed])

    xlsx_path = outdir / f"{handle} - Recovered Tweets.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


def _tweet_row(r):
    return [r["date"], r["text"], r["date_confidence"], r["source"],
            r["archive_url"], r["tweet_id"], r["capture_timestamp"],
            r.get("recovered_via", "status"), str(r.get("is_retweet", False)),
            r.get("retweeted_user") or ""]


def _failed_row(r):
    return [r["tweet_id"], r["reason"], str(r["needs_browser"]),
            r["attempted_captures"], r["original_url"]]


def _write_table(ws, cols, rows):
    ws.append([c for c, _ in cols])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, (_, width) in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"


def _write_summary(ws, handle, recovered, failed, raw_captures, unique_ids,
                   page_stats=None, pages_done=None, manifest_ids=None):
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 28

    fail_reasons = Counter(r["reason"] for r in failed)
    conf = Counter(r["date_confidence"] for r in recovered)
    hosts = Counter(r["source"] for r in recovered)
    host_line = " · ".join(f"{h}: {n}" for h, n in hosts.most_common())
    dates = [r["date"][:10] for r in recovered]

    # Recovery rate over the account's own tweets: retweets are separate
    # (they aren't part of the per-ID universe), and timeline mining can
    # surface own-tweet IDs the status-URL discovery never saw — count
    # those in the denominator so the rate stays honest.
    own = [r for r in recovered if not r.get("is_retweet", False)]
    retweets = len(recovered) - len(own)
    via = Counter(r.get("recovered_via", "status") for r in own)
    extra_ids = 0
    if manifest_ids is not None:
        known = {r["tweet_id"] for r in own} | {r["tweet_id"] for r in failed}
        extra_ids = len(known - manifest_ids)
    denom = unique_ids + extra_ids
    rate = f"{len(own) / denom:.1%}" if denom else "—"

    rows = [
        (f"{handle} — Recovered Tweets", None),
        ("Source: Internet Archive Wayback Machine", None),
        (f"Generated {datetime.now():%Y-%m-%d}", None),
        (None, None),
        ("Unique tweet IDs found in archive", denom),
        ("   — from individual tweet URLs", unique_ids),
        ("   — only seen on profile-page captures", extra_ids),
        ("Own tweets recovered (text + date)", len(own)),
        ("   — via individual tweet captures", via.get("status", 0)),
        ("   — via profile-page captures", via.get("timeline", 0)),
        ("Retweets captured (from profile pages)", retweets),
        ("Recovery rate (own tweets)", rate),
        ("Dates — exact", conf.get("exact", 0)),
        ("Dates — approximate", conf.get("approximate", 0)),
        ("Failed (no recoverable text)", len(failed)),
    ]
    rows += [(f"   — {reason}", n) for reason, n in fail_reasons.most_common()]
    rows += [
        ("Date range covered", f"{min(dates)} → {max(dates)}" if dates else "—"),
        ("Raw captures examined", raw_captures),
    ]
    if page_stats:
        page_status = Counter(p["status"] for p in (pages_done or []))
        rows += [
            ("Profile-page captures found", page_stats.get("raw_page_captures", 0)),
            ("Profile-page captures selected", page_stats.get("selected", 0)),
            ("   — parsed", page_status.get("parsed", 0)),
            ("   — JS shell (unrendered)", page_status.get("js_shell", 0)),
        ]
    rows += [
        ("Source hosts", host_line or "—"),
        (None, None),
        ("Method / caveats", None),
        (METHOD_NOTE, None),
    ]
    for label, value in rows:
        ws.append([label, value])
    ws["A1"].font = Font(bold=True, size=14)
    ws.cell(row=len(rows), column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[len(rows)].height = 90
